from openpyxl import load_workbook  # Импортируем всё необходимое.
from openpyxl import Workbook

compare_set_1 = set()  # Создаём изначальный список.
compare_set_2 = set()  # Создаём список для сравнения.
dbwb = load_workbook(filename='All_DEALS.xlsx')  # Загружаем текущую базу.
dbws = dbwb['Лист1']
for i in dbws['A']:  # Добавляем базу в изначальный список.
    compare_set_1.add(i.value)

b_swb = load_workbook(filename=r'C:\Users\Admin\Desktop\Источники данных\Аналитика 2.0\Битрикс.xlsx')  # Загружаем
# свежий Битрикс.
b_sws = b_swb['Битрикс']
for i in b_sws['A']:  # Добавляем Битрикс в список для сравнения.
    compare_set_2.add(i.value)
compare_set_2.discard('ID')  # Убираем "ID".

for i in compare_set_2:  # Проверяем оставшиеся позиции второго списка на тип данных.
    if type(i) != int:
        i = int(i)

lost_deals = compare_set_1 - compare_set_2  # Старая база - новая база = потеряшки.
while None in lost_deals:
    lost_deals.remove(None)

new_deals = compare_set_2 - compare_set_1 # Новая база - старые сделки = новые сделки.
while None in new_deals:
    new_deals.remove(None)

print('Новые сделки: ', len(new_deals))  # Вывод количества новых сделок.
print()  # Разрыв строки.
print('Потерянные сделки: ', len(lost_deals))  # Вывод количества потерянных сделок.

for i in new_deals:  # Добавляем к основной базе новые сделки в колоночку.
    dbws.append([i, ])
dbwb.save('All_DEALS.xlsx')  # Сохраняем базу данных.

lost_wb = Workbook()  # Создаём рабочую книгу под потеряшек.
lost_ws = lost_wb.active
for i in lost_deals:  # Добавляем каждую потеряшку в таблицу с новой строки.
    lost_ws.append([i, ])
lost_wb.save('LOST.xlsx')  # Сохраняем потеряшек.
